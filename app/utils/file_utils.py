import os
import json
import pytesseract
from PyPDF2 import PdfReader
from pdf2image import convert_from_path
from PIL import Image, ImageOps
from docx import Document
import csv
from io import StringIO
from zipfile import ZipFile
from openpyxl import load_workbook

async def extract_text(file_path: str) -> str:
    """
    Extract text from various file types: PDF, images, DOCX, XLSX, CSV, JSON, TXT, ZIP.
    - PDF: try native text, fallback to OCR if needed
    - Image: OCR
    - DOCX: python-docx
    - XLSX: openpyxl
    - CSV: csv.reader
    - JSON: json.dumps pretty
    - ZIP: list filenames + contents (if text-based)
    - TXT: read as plain text
    """
    ext = os.path.splitext(file_path)[1].lower()
    text = ""
    print(f"[DEBUG] Extracting text from: {file_path} (ext: {ext})")

    # --------------- PDF ---------------
    if ext == ".pdf":
        reader = PdfReader(file_path)
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text() or ""
            print(f"[DEBUG] PDF page {i} text length: {len(page_text)}")

            # OCR fallback jika kosong / terlalu pendek
            if not page_text.strip() or len(page_text.strip()) < 20:
                print(f"[DEBUG] Page {i} empty, applying OCR fallback")
                images = convert_from_path(file_path, first_page=i+1, last_page=i+1)
                for img in images:
                    img = img.convert("L")
                    img = ImageOps.invert(img)
                    ocr_text = pytesseract.image_to_string(img)
                    print(f"[DEBUG] OCR page {i} text length: {len(ocr_text)}")
                    page_text += ocr_text

            text += page_text

    # --------------- Images ---------------
    elif ext in [".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".webp"]:
        img = Image.open(file_path).convert("L")
        img = ImageOps.invert(img)
        text = pytesseract.image_to_string(img)
        print(f"[DEBUG] Image OCR text length: {len(text)}")

    # --------------- DOCX ---------------
    elif ext == ".docx":
        try:
            doc = Document(file_path)
            paragraphs = [para.text for para in doc.paragraphs]
            text = "\n".join(paragraphs)
            print(f"[DEBUG] DOCX paragraphs extracted: {len(paragraphs)}")
        except Exception as e:
            print(f"[ERROR] Failed to parse DOCX: {e}")

    # --------------- XLSX ---------------
    elif ext in [".xlsx", ".xls"]:
        try:
            import openpyxl
            from xlrd import open_workbook

            if ext == ".xlsx":
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"

            elif ext == ".xls":
                book = open_workbook(file_path)
                for sheet in book.sheets():
                    for row_idx in range(sheet.nrows):
                        row = sheet.row_values(row_idx)
                        text += " ".join([str(cell) for cell in row if cell]) + "\n"

            print(f"[DEBUG] Excel text length: {len(text)}")

        except Exception as e:
            print(f"[ERROR] Failed to parse Excel file: {e}")

    # --------------- CSV ---------------
    elif ext == ".csv":
        try:
            with open(file_path, newline='', encoding='utf-8', errors='ignore') as csvfile:
                reader = csv.reader(csvfile)
                for row in reader:
                    text += "\t".join(row) + "\n"
            print(f"[DEBUG] CSV text length: {len(text)}")
        except Exception as e:
            print(f"[ERROR] Failed to parse CSV: {e}")

    # --------------- JSON ---------------
    elif ext == ".json":
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            text = json.dumps(data, indent=2, ensure_ascii=False)
            print(f"[DEBUG] JSON text length: {len(text)}")
        except Exception as e:
            print(f"[ERROR] Failed to parse JSON: {e}")

    # --------------- ZIP (Optional: ekstrak semua text file di dalamnya) ---------------
    elif ext == ".zip":
        try:
            with ZipFile(file_path, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                text += f"--- ZIP Contents ---\n" + "\n".join(file_list) + "\n"
                for name in file_list:
                    if name.endswith((".txt", ".csv", ".json")):
                        with zip_ref.open(name) as f:
                            try:
                                content = f.read().decode("utf-8", errors="ignore")
                                text += f"\n--- File: {name} ---\n{content}\n"
                            except Exception as e:
                                print(f"[WARN] Failed to read {name} from zip: {e}")
            print(f"[DEBUG] ZIP text length: {len(text)}")
        except Exception as e:
            print(f"[ERROR] Failed to parse ZIP: {e}")

    # --------------- Plain Text ---------------
    else:
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
            print(f"[DEBUG] Plain text length: {len(text)}")
            print(f"[DEBUG] Plain text preview: {text[:200]}")
        except Exception as e:
            print(f"[ERROR] Failed to read as plain text: {e}")

    return text
